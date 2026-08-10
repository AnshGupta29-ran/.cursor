import json
import os
from pathlib import Path
from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.responses import JSONResponse
from pydantic import BaseModel
import pandas as pd
import io

# Directories
DATA_DIR = Path(__file__).parent / "data"
MODEL_DIR = DATA_DIR / "models"
OUTPUT_DIR = Path(__file__).parent / "output"
LOG_FILE = Path(__file__).parent / "logs" / "holdfast.log"

DATA_DIR.mkdir(exist_ok=True)
MODEL_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)
(Path(__file__).parent / "logs").mkdir(exist_ok=True)

app = FastAPI(title="Holdfast Retention Workbench")

# Simple in‑memory config (thresholds) – persisted in workbook on next run
DEFAULT_THRESHOLDS = {"red": 0.7, "amber": 0.4}

def log_event(event: str, detail: dict | None = None):
    entry = {
        "timestamp": pd.Timestamp.utcnow().isoformat(),
        "event": event,
        "detail": detail or {},
    }
    with LOG_FILE.open("a", encoding="utf-8") as f:
        f.write(json.dumps(entry) + "\n")

def model_path():
    return MODEL_DIR / "model.joblib"

def meta_path():
    return MODEL_DIR / "model_meta.json"

def load_model():
    if not model_path().exists():
        raise FileNotFoundError("No trained model")
    from joblib import load
    return load(str(model_path()))

def save_model(model, meta: dict):
    from joblib import dump
    dump(model, str(model_path()))
    with meta_path().open("w", encoding="utf-8") as f:
        json.dump(meta, f)

def register_run(record: dict):
    run_path = DATA_DIR / "run.json"
    with run_path.open("w", encoding="utf-8") as f:
        json.dump(record, f, indent=2)

def generate_workbook(run_record: dict, scores: pd.DataFrame | None = None, importances: pd.Series | None = None):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    wb = Workbook()
    # 1. Start_Here
    ws = wb.active
    ws.title = "Start_Here"
    ws["A1"] = "Holdfast Retention Workbench"
    ws["A2"] = f"Last run: {run_record.get('timestamp', '')}"
    ws["A3"] = "Paste new member rows into the 'Score_Input' sheet and POST /jobs/score."
    # 2. Config
    ws = wb.create_sheet("Config")
    ws["A1"] = "Model"
    ws["B1"] = run_record.get("model_type", "")
    ws["A2"] = "Red threshold"
    ws["B2"] = DEFAULT_THRESHOLDS["red"]
    ws["A3"] = "Amber threshold"
    ws["B3"] = DEFAULT_THRESHOLDS["amber"]
    # 3. Score_Input (header only)
    ws = wb.create_sheet("Score_Input")
    ws.append(["member_id", "tenure_months", "fob_entries_30d", "shop_nights_90d", "tool_checkouts_90d", "orientation_completed", "dues_current", "lapsed"])
    # 4. Risk_Register
    ws = wb.create_sheet("Risk_Register")
    ws.append(["member_id", "probability", "tier", "top_3_reasons"])
    if scores is not None:
        for _, row in scores.iterrows():
            ws.append([row["member_id"], row["probability"], row["tier"], ", ".join(row["reasons"])])
    # simple tier coloring
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    amber_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    for r in range(2, ws.max_row + 1):
        tier = ws[f"C{r}"].value
        if tier == "Red":
            for c in range(1, ws.max_column + 1):
                ws[f"{chr(64 + c)}{r}"].fill = red_fill
        elif tier == "Amber":
            for c in range(1, ws.max_column + 1):
                ws[f"{chr(64 + c)}{r}"].fill = amber_fill
    # 5. Feature_Importances
    ws = wb.create_sheet("Feature_Importances")
    ws.append(["feature", "importance"])
    if importances is not None:
        for feat, imp in importances.items():
            ws.append([feat, float(imp)])
    # 6. Run_History (single entry)
    ws = wb.create_sheet("Run_History")
    ws.append(["run_id", "timestamp", "rows", "metrics", "status"])
    ws.append([
        run_record.get("run_id", ""),
        run_record.get("timestamp", ""),
        run_record.get("row_count", ""),
        json.dumps(run_record.get("metrics", {})),
        run_record.get("status", ""),
    ])
    out_path = OUTPUT_DIR / "holdfast_workbook.xlsx"
    wb.save(str(out_path))
    log_event("workbook_generated", {"path": str(out_path)})
    return out_path

def validate_csv(df: pd.DataFrame, required_cols: list[str]):
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")
    # numeric coercion
    num_cols = [c for c in required_cols if c != "member_id" and df[c].dtype.kind in "biufc"]
    for col in num_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    return df

class TrainResponse(BaseModel):
    job_id: str
    status: str
    metrics: dict
    model_type: str

class ScoreResponse(BaseModel):
    job_id: str
    status: str
    results_path: str

@app.post("/jobs/train", response_model=TrainResponse)
async def train_job(file: UploadFile = File(...)):
    # Skipping content_type check, accepting any file type
    content = await file.read()
    df = pd.read_csv(io.StringIO(content.decode()))
    required = ["member_id", "tenure_months", "fob_entries_30d", "shop_nights_90d", "tool_checkouts_90d", "orientation_completed", "dues_current", "lapsed"]
    try:
        df = validate_csv(df, required)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    if len(df) < 40:
        raise HTTPException(status_code=400, detail="Need at least 40 rows for training")
    if df["lapsed"].nunique() < 2:
        raise HTTPException(status_code=400, detail="Both churned and non‑churned examples are required")
    # simple pipeline
    from sklearn.compose import ColumnTransformer
    from sklearn.preprocessing import OneHotEncoder, StandardScaler
    from sklearn.impute import SimpleImputer
    from sklearn.pipeline import Pipeline
    from sklearn.linear_model import LogisticRegression
    from sklearn.metrics import accuracy_score, precision_score, recall_score, roc_auc_score
    X = df.drop(columns=["member_id", "lapsed"])
    y = df["lapsed"]
    numeric_features = X.select_dtypes(include=["int64", "float64"]).columns.tolist()
    categorical_features = []  # none for this demo
    preprocessor = ColumnTransformer(
        transformers=[
            ("num", Pipeline([("imputer", SimpleImputer(strategy="median")), ("scaler", StandardScaler())]), numeric_features),
        ]
    )
    clf = Pipeline(steps=[("preprocessor", preprocessor), ("classifier", LogisticRegression(random_state=42, max_iter=200))])
    from sklearn.model_selection import train_test_split
    X_train, X_val, y_train, y_val = train_test_split(X, y, test_size=0.2, stratify=y, random_state=42)
    clf.fit(X_train, y_train)
    preds = clf.predict(X_val)
    probs = clf.predict_proba(X_val)[:, 1]
    metrics = {
        "accuracy": accuracy_score(y_val, preds),
        "precision": precision_score(y_val, preds, zero_division=0),
        "recall": recall_score(y_val, preds, zero_division=0),
        "roc_auc": roc_auc_score(y_val, probs),
    }
    # save model
    save_model(clf, {"model_type": "LogisticRegression", "metrics": metrics})
    run_record = {
        "run_id": "run-1",
        "timestamp": pd.Timestamp.utcnow().isoformat(),
        "row_count": len(df),
        "metrics": metrics,
        "status": "succeeded",
        "model_type": "LogisticRegression",
    }
    register_run(run_record)
    # generate empty workbook (no scores yet)
    generate_workbook(run_record)
    log_event("train_completed", {"metrics": metrics})
    return TrainResponse(job_id="train-1", status="succeeded", metrics=metrics, model_type="LogisticRegression")

def tier_from_prob(p: float, thresholds: dict):
    if p >= thresholds["red"]:
        return "Red"
    if p >= thresholds["amber"]:
        return "Amber"
    return "Green"

@app.post("/jobs/score", response_model=ScoreResponse)
async def score_job(file: UploadFile = File(...)):
    if not model_path().exists():
        raise HTTPException(status_code=409, detail="No trained model available – run training first.")
    # Skipping content_type check, accepting any file type
    content = await file.read()
    df = pd.read_csv(io.StringIO(content.decode()))
    required = ["member_id", "tenure_months", "fob_entries_30d", "shop_nights_90d", "tool_checkouts_90d", "orientation_completed", "dues_current"]
    try:
        df = validate_csv(df, required)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    model = load_model()
    X = df.drop(columns=["member_id"])
    probs = model.predict_proba(X)[:, 1]
    prob_series = pd.Series(probs)
    df["probability"] = prob_series
    df["tier"] = prob_series.apply(lambda p: tier_from_prob(p, DEFAULT_THRESHOLDS))
    # simple reasons: top 3 features with highest absolute coefficient * value
    # for demo, generate placeholder reasons
    def reasons(row):
        top = []
        # placeholder strings
        if row["fob_entries_30d"] < 5:
            top.append("very few fob entries in last 30 days")
        if row["shop_nights_90d"] == 0:
            top.append("no shop nights in 90 days")
        if row["dues_current"] == 0:
            top.append("dues not current")
        return top[:3]
    df["reasons"] = df.apply(reasons, axis=1)
    # load last run record for workbook generation
    run_path = DATA_DIR / "run.json"
    run_record = json.loads(run_path.read_text()) if run_path.exists() else {}
    # update run record with metrics placeholder
    generate_workbook(run_record, scores=df, importances=None)
    out_path = OUTPUT_DIR / "holdfast_workbook.xlsx"
    log_event("score_completed", {"rows": len(df)})
    return ScoreResponse(job_id="score-1", status="succeeded", results_path=str(out_path))

@app.get("/healthz")
async def health():
    return {"status": "ok"}

@app.get("/metrics")
async def metrics():
    # simple counters from log file
    cnt = {"train": 0, "score": 0}
    if LOG_FILE.exists():
        for line in LOG_FILE.read_text().splitlines():
            try:
                entry = json.loads(line)
                ev = entry.get("event")
                if ev == "train_completed":
                    cnt["train"] += 1
                if ev == "score_completed":
                    cnt["score"] += 1
            except Exception:
                continue
    return cnt
